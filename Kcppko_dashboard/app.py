# app.py
import streamlit as st
import pandas as pd
import io
from datetime import datetime
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

st.set_page_config(
    page_title="PKO Kernel Dashboard",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ===========================
# Sidebar - Settings & Load
# ===========================
st.sidebar.title("Settings & Data")
PKO_rate = st.sidebar.number_input("PKO rate (fraction)", value=0.46, format="%.4f")
PKC_rate = st.sidebar.number_input("PKC rate (fraction)", value=0.54, format="%.4f")
company_name = st.sidebar.text_input("Company name", value="Nama Perusahaan")
default_ffa = st.sidebar.number_input("Default FFA (%)", value=2.5, format="%.3f")
default_moist = st.sidebar.number_input("Default Moisture (%)", value=0.2, format="%.3f")

st.sidebar.markdown("---")
st.sidebar.markdown("**Load / Save**")
uploaded = st.sidebar.file_uploader("Upload CSV (historical data)", type=["csv"])
if uploaded:
    try:
        df_loaded = pd.read_csv(uploaded, parse_dates=["Tanggal"])
        # normalize column names if needed
        st.session_state.df = df_loaded.copy()
        st.sidebar.success("CSV loaded into session.")
    except Exception as e:
        st.sidebar.error(f"Failed to load CSV: {e}")

# ===========================
# Initialize session storage
# ===========================
if "df" not in st.session_state:
    st.session_state.df = pd.DataFrame(columns=[
        "Tanggal", "Kernel_ton",
        "FFA_Pagi", "FFA_Siang", "FFA_Malam", "FFA_Harian",
        "Moist_Pagi", "Moist_Siang", "Moist_Malam", "Moist_Harian",
        "PKO_ton", "PKC_ton", "Losses_ton", "Rendemen_PKO_pct", "Rendemen_PKC_pct",
        "Catatan"
    ])

# ===========================
# Input form
# ===========================
st.title("🌴 PKO Kernel Processing — Dashboard")
st.markdown("Input data harian dan otomatiskan laporan PKO/PKC, kualitas (FFA/Moist), dan grafik.")

with st.form("entry_form", clear_on_submit=False):
    col1, col2 = st.columns([2,1])
    with col1:
        tanggal = st.date_input("Tanggal", value=datetime.today().date())
        kernel = st.number_input("Kernel diproses (ton)", min_value=0.0, value=0.0, step=0.1, format="%.2f")
    with col2:
        note = st.text_input("Catatan (opsional)")
    st.markdown("**FFA (%) — 3 shift**")
    c1, c2, c3 = st.columns(3)
    with c1:
        ffa_pagi = st.number_input("FFA Pagi", value=float(default_ffa), format="%.3f", key="ffa_pagi")
    with c2:
        ffa_siang = st.number_input("FFA Siang", value=float(default_ffa), format="%.3f", key="ffa_siang")
    with c3:
        ffa_malam = st.number_input("FFA Malam", value=float(default_ffa), format="%.3f", key="ffa_malam")

    st.markdown("**Moisture (%) — 3 shift**")
    m1, m2, m3 = st.columns(3)
    with m1:
        moist_pagi = st.number_input("Moist Pagi", value=float(default_moist), format="%.3f", key="moist_pagi")
    with m2:
        moist_siang = st.number_input("Moist Siang", value=float(default_moist), format="%.3f", key="moist_siang")
    with m3:
        moist_malam = st.number_input("Moist Malam", value=float(default_moist), format="%.3f", key="moist_malam")

    submitted = st.form_submit_button("Simpan Data Harian")

def compute_row(kernel, ffa_vals, moist_vals, pko_rate, pkc_rate):
    # average of provided shift values
    ffa_h = round(sum(ffa_vals) / len(ffa_vals), 3)
    moist_h = round(sum(moist_vals) / len(moist_vals), 3)
    pko = round(kernel * pko_rate, 3)
    pkc = round(kernel * pkc_rate, 3)
    losses = round(kernel - (pko + pkc), 3)
    rend_pko = round((pko / kernel * 100) if kernel > 0 else 0, 3)
    rend_pkc = round((pkc / kernel * 100) if kernel > 0 else 0, 3)
    return ffa_h, moist_h, pko, pkc, losses, rend_pko, rend_pkc

if submitted:
    ffa_vals = [ffa_pagi, ffa_siang, ffa_malam]
    moist_vals = [moist_pagi, moist_siang, moist_malam]
    ffa_h, moist_h, pko, pkc, losses, rend_pko, rend_pkc = compute_row(kernel, ffa_vals, moist_vals, PKO_rate, PKC_rate)
    row = {
        "Tanggal": pd.to_datetime(tanggal).strftime("%Y-%m-%d"),
        "Kernel_ton": kernel,
        "FFA_Pagi": ffa_pagi, "FFA_Siang": ffa_siang, "FFA_Malam": ffa_malam, "FFA_Harian": ffa_h,
        "Moist_Pagi": moist_pagi, "Moist_Siang": moist_siang, "Moist_Malam": moist_malam, "Moist_Harian": moist_h,
        "PKO_ton": pko, "PKC_ton": pkc, "Losses_ton": losses,
        "Rendemen_PKO_pct": rend_pko, "Rendemen_PKC_pct": rend_pkc,
        "Catatan": note
    }
    st.session_state.df = pd.concat([st.session_state.df, pd.DataFrame([row])], ignore_index=True)
    st.success("Data tersimpan di session. Gunakan Export untuk menyimpan permanen.")

# ===========================
# Data table + actions
# ===========================
st.subheader("📋 Data Harian (Session)")
st.dataframe(st.session_state.df, use_container_width=True)

col_a, col_b, col_c = st.columns([1,1,1])
with col_a:
    if st.button("Export CSV"):
        csv_bytes = st.session_state.df.to_csv(index=False).encode("utf-8")
        st.download_button("Download CSV", data=csv_bytes, file_name=f"PKO_data_{datetime.now().strftime('%Y%m%d')}.csv", mime="text/csv")
with col_b:
    if st.button("Export Excel (.xlsx)"):
        bio = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        # write header + rows
        for r in [st.session_state.df.columns.tolist()] + st.session_state.df.values.tolist():
            ws.append(r)
        # formatting header bold
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        wb.save(bio)
        bio.seek(0)
        st.download_button("Download XLSX", data=bio, file_name=f"PKO_data_{datetime.now().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
with col_c:
    if st.button("Clear Session"):
        st.session_state.df = st.session_state.df.iloc[0:0]
        st.success("Session cleared.")

# ===========================
# Dashboard: charts & KPIs
# ===========================
st.subheader("📈 Dashboard & Trends")
df = st.session_state.df.copy()
if not df.empty:
    df["Tanggal_dt"] = pd.to_datetime(df["Tanggal"])
    df = df.sort_values("Tanggal_dt")
    # KPI row
    kpi_col1, kpi_col2, kpi_col3, kpi_col4 = st.columns(4)
    total_kernel = df["Kernel_ton"].sum()
    total_pko = df["PKO_ton"].sum()
    avg_ffa = df["FFA_Harian"].mean()
    avg_moist = df["Moist_Harian"].mean()
    kpi_col1.metric("Total Kernel (ton)", f"{total_kernel:,.2f}")
    kpi_col2.metric("Total PKO (ton)", f"{total_pko:,.2f}")
    kpi_col3.metric("Avg FFA (%)", f"{avg_ffa:.3f}")
    kpi_col4.metric("Avg Moist (%)", f"{avg_moist:.3f}")

    # Production line
    fig_prod = px.line(df, x="Tanggal_dt", y=["PKO_ton","PKC_ton"], markers=True, labels={"value":"Ton","Tanggal_dt":"Tanggal"})
    fig_prod.update_layout(title="Produksi PKO & PKC")
    st.plotly_chart(fig_prod, use_container_width=True)

    # Quality trend
    fig_q = px.line(df, x="Tanggal_dt", y=["FFA_Harian","Moist_Harian"], markers=True, labels={"value":"%","Tanggal_dt":"Tanggal"})
    fig_q.update_layout(title="Trend Kualitas - FFA & Moisture")
    st.plotly_chart(fig_q, use_container_width=True)

    # Rendemen grouped bar
    fig_r = px.bar(df, x="Tanggal_dt", y=["Rendemen_PKO_pct","Rendemen_PKC_pct"], barmode="group", labels={"value":"%","Tanggal_dt":"Tanggal"}, title="Rendemen (%) per Hari")
    st.plotly_chart(fig_r, use_container_width=True)
else:
    st.info("Belum ada data. Masukkan data harian di form atas.")

# ===========================
# Ready-to-print single day report
# ===========================
st.subheader("🖨️ Siap Cetak / Laporan Harian")
with st.expander("Cetak Laporan Harian (pilih tanggal)"):
    if df.empty:
        st.write("Data kosong.")
    else:
        options = df["Tanggal"].tolist()
        pick = st.selectbox("Pilih tanggal", options=options)
        row = df[df["Tanggal"] == pick].iloc[0]
        # present as two-column layout
        left, right = st.columns([2,1])
        with left:
            st.markdown(f"### {company_name}")
            st.markdown(f"**Laporan Harian** — Tanggal: **{pick}**")
            st.write("")
            st.table(pd.DataFrame(row).rename(columns={0:"Value"}))
        with right:
            st.markdown("**Tanda Terima / Persetujuan**")
            st.write("Operator: ___________________")
            st.write("Supervisor: __________________")
            st.write("Manager: ____________________")
        # export single report xlsx
        if st.button("Export Laporan Terpilih ke XLSX"):
            bio = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan Harian"
            ws.append(["Field","Value"])
            for k,v in row.items():
                ws.append([k, v])
            # format
            for cell in ws[1]:
                cell.font = Font(bold=True)
            wb.save(bio)
            bio.seek(0)
            st.download_button("Download Laporan XLSX", data=bio, file_name=f"Laporan_{pick}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Footer note
st.markdown("---")
st.caption("Streamlit app untuk laporan PKO Kernel. Export ke XLSX/CSV untuk simpan permanen dan cetak (print to PDF).")
